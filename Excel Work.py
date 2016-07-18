from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string as ColToInt
from time import strftime

date = strftime('%B %d %Y')  # Gets date in the from of MONTH DATE YEAR
saveFile = 'ATT.xlsx'  # Sets the save file
wb = load_workbook(saveFile)
ws = wb.active


def save():  # saves file
    wb.save(saveFile)


def find(findValue, row=None, column=None):
    if isinstance(findValue, str):
        findValue = findValue.lower()
    if row is None and column is None:
        for row in ws.iter_rows:
            for cell in row:
                if cell.value is findValue:
                    return cell.row, ColToInt(cell.column)

    elif row is not None and column is None:  # scanning through a row
        index = 1
        while True:
            cell = ws.cell(row=row, column=index)
            if cell.value is findValue:
                return cell.row, ColToInt(cell.column)
            index += 1

    elif row is None and column is not None:  # scanning through a column
        index = 2
        lastCell = ws.cell(row=index - 1, column=column)
        while True:
            cell = ws.cell(row=index, column=column)
            if lastCell.value is None and cell.value is None:
                return None, None
            if cell.value == findValue:
                return cell.row, ColToInt(cell.column)
            index += 1
            lastCell = cell


def create_attendance_date(row, column):  # Creates meeting date
    column = ColToInt(column)  # Converts letter column to number
    ws.cell(row=row, column=column, value=date)  # Adds date column
    save()  # Saves file


def find_empty_date_column():
    global attCol
    index = 2  # starts column count at 2('B')
    lastCell = ws['A1']  # first cell is set as 'A1'
    while True:  # runs loop to find empty cell in row
        cell = ws.cell(row=1, column=index)
        if lastCell.value is None and cell.value is None:
            # [date][None][None]...[None] finds 2 empty cells and the
            # uses the previous cell as the point at which to add the
            # date stamp
            create_attendance_date(lastCell.row, lastCell.column)
            attCol = (cell.column)
            break

        elif cell.value == date:

            # Makes attCol known to the whole program for attendance registration
            attCol = (cell.column)
            break

        # elif cell.value is None and lastCell.value is not None:
        index += 1
        lastCell = cell


def emptyRow():
    return find(None, None, 1)[0]


def number_check_in(num):
    timeIn = strftime("%I:%M %p")
    row, col = find(num, None, 1)
    if row is None and col is None:
        add_member()
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        timeIn = strftime("%I:%M %p")
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
    save()


def name_check_in(name):
    row, col = find(name, None, 2)
    if row is None and col is None:
        add_member()
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        timeIn = strftime("%I:%M %p")
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
    save()


def check_in(number=None, name=None):
    if number is not None:
        # Name based check in
        number_check_in(number)
    else:
        # Number based check in
        name_check_in(name.lower())


def approve_payment(identifier):
    if isinstance(identifier, str):
        identifier = identifier.lower()
        coll = 2
    else:
        coll = 1
    row, col = find(identifier, None, coll)
    if ws.cell(row=row, column=3).value is None:
        ws.cell(row=row, column=3, value='PAID')
    save()


def add_member():
    Fn = input('Enter First Name: ')
    Ln = input('Enter Last Name: ')
    FLname = Fn + Ln
    FLname = FLname.lower()
    number = int(input('Number: '))
    row = emptyRow()
    ws.cell(row=row, column=1, value=number)
    ws.cell(row=row, column=2, value=FLname)
    save()


find_empty_date_column()
name_check_in('xc')
# number_check_in(99)
# approve_payment('js')
