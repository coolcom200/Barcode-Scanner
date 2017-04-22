import random
from openpyxl import load_workbook
import string

saveFile = 'Attendance.xlsx'  # Sets the save file
wb = load_workbook(saveFile)
ws = wb.active
array = []


def id_generator(size=6, chars=string.ascii_lowercase):
    x = ''.join(random.choice(chars) for _ in range(size))
    x = x[:3] + " " + x[3:]
    return x


for i in range(2, 100):
    x = random.randint(100000000, 999999999)
    array.append(str(x))
    # ws.cell(row=i, column=1, value=x)
    ws.cell(row=i, column=2, value=id_generator())
print(array)

wb.save(saveFile)

