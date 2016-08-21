from tkinter import *
import threading
import time
from tkinter import messagebox
from sys import exit as Stp
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string as ColToInt
from matplotlib import pyplot as plt

date = time.strftime('%B %d %Y')  # Gets date in the from of MONTH DATE YEAR
saveFile = 'ATT.xlsx'  # Sets the save file
wb = load_workbook(saveFile)
ws = wb.active
counter = False

def save():  # saves file
    wb.save(saveFile)


def find(findValue, row=None, column=None):
    if isinstance(findValue, str):
        findValue = findValue.lower()
    if row is None and column is None:
        for row in ws.iter_rows:
            for cell in row:
                if cell.value is findValue:
                    return cell.row, ColToInt(cell.column)

    elif row is not None and column is None:  # scanning through a row
        index = 1
        while True:
            cell = ws.cell(row=row, column=index)
            if cell.value is findValue:
                return cell.row, ColToInt(cell.column)
            index += 1

    elif row is None and column is not None:  # scanning through a column
        index = 2
        lastCell = ws.cell(row=index - 1, column=column)
        while True:
            cell = ws.cell(row=index, column=column)
            if lastCell.value is None and cell.value is None:
                return None, None
            if cell.value == findValue:
                return cell.row, ColToInt(cell.column)
            index += 1
            lastCell = cell


def create_attendance_date(row, column):  # Creates meeting date
    column = ColToInt(column)  # Converts letter column to number
    ws.cell(row=row, column=column, value=date)  # Adds date column
    save()  # Saves file


def find_empty_date_column():
    global attCol
    index = 2  # starts column count at 2('B')
    lastCell = ws['A1']  # first cell is set as 'A1'
    while True:  # runs loop to find empty cell in row
        cell = ws.cell(row=1, column=index)
        if lastCell.value is None and cell.value is None:
            # [date][None][None]...[None] finds 2 empty cells and the
            # uses the previous cell as the point at which to add the
            # date stamp
            create_attendance_date(lastCell.row, lastCell.column)
            attCol = lastCell.column
            break

        elif cell.value == date:

            # Makes attCol known to the whole program for attendance registration
            attCol = cell.column
            break

        # elif cell.value is None and lastCell.value is not None:
        index += 1
        lastCell = cell


def emptyRow():
    return find(None, None, 1)[0]


def number_check_in(num):
    timeIn = time.strftime("%I:%M %p")
    row, col = find(num, None, 1)
    if row is None and col is None:
        add_member(num)
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
        add_name_listbox(row)
    save()


def name_check_in(name):
    row, col = find(name, None, 2)
    if row is None and col is None:
        add_member(name, True)
    elif ws.cell(row=row, column=ColToInt(attCol)).value is None:
        timeIn = time.strftime("%I:%M %p")
        ws.cell(row=row, column=ColToInt(attCol), value=timeIn)
        add_name_listbox(row)
    save()


def check_in(identifier, UserName):
    find_empty_date_column()
    if UserName:
        # Name based check in
        popup.destroy()
        name_check_in(identifier.lower())
    else:
        # Number based check in
        number_check_in(identifier)


def approve_payment(identifier):
    if isinstance(identifier, str):
        identifier = identifier.lower()
        coll = 2
    else:
        coll = 1
    row, col = find(identifier, None, coll)
    if ws.cell(row=row, column=3).value is None:
        ws.cell(row=row, column=3, value='PAID')
    save()


def gather_att_data():
    indices = []
    names = []
    attendData = []
    colTup = ws.columns
    for column in colTup:
        present = 0
        for cell in column:
            if cell.value is not None:
                present += 1

        attendData.append(present - 1)
    attendData = attendData[3:]
    for i in range(len(attendData)):
        indices.append(i)
        if i % 5 == 0:
            names.append('Mt #' + str(i + 1))
        else:
            names.append('')
    return attendData, indices, names


def show_graph():
    data, indices, label = gather_att_data()
    plt.bar(indices, data, 0.6, color='green', edgecolor='green')
    indices = [x + 0.3 for x in indices]
    plt.xticks(indices, label)
    plt.ylim([0, max(data) + 1])
    plt.show()


def add_member(identifier, strr=False):
    choice = messagebox.askquestion('Not In Database',
                                    'Sorry you are not in the database.\nWould you like to become a member?')

    def create_mem_row(num, name):
        num = int(num)
        row = emptyRow()
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=name)
        check_in(num, False)
        save()

    if choice == 'yes':
        FLname = None
        StuNumber = None
        questionF = Toplevel()
        if strr:
            FLname = identifier
            Label(questionF, text='Enter Student Number:').pack()
            num = Entry(questionF)
            num.pack()

            def num_ret():
                StuNumber = num.get()
                questionF.destroy()
                create_mem_row(StuNumber, FLname)

            Button(questionF, text='Submit', font='Times 14 bold', relief="groove", command=num_ret).pack()

        else:
            StuNumber = identifier
            Label(questionF, text='Enter First Name: ').pack()
            Fn = Entry(questionF)
            Fn.pack()
            Label(questionF, text='Enter Last Name: ').pack()
            Ln = Entry(questionF)
            Ln.pack()

            def name_ret():
                FLname = Fn.get() + ' ' + Ln.get()
                FLname = FLname.lower()
                create_mem_row(StuNumber, FLname)
                questionF.destroy()

            Button(questionF, text='Submit', font='Times 14 bold', relief="groove", command=name_ret).pack()


def check_eligible(ID):
    row = find(ID, None, 2)[0]
    days = ws.iter_rows('D' + str(row) + ':' + attCol + str(row))
    notPres = 0
    attend = 0
    for i in list(days)[0]:
        if i.value is None:
            notPres += 1
        else:
            attend += 1
    perc = round(attend / (attend + notPres) * 100, 0)

    # must add math to count attendance to classes
    if perc > 60:
        return 'Yes'
    else:
        return 'No'


signInList = []

root = Tk()
root.resizable(width=False, height=False)


def name_check_in_GUI():
    global popup, FN, LN
    popup = Toplevel()
    popup.tkraise(root)
    Label(popup, text='Enter First Name').grid(row=1, column=1, sticky=W)
    FN = Entry(popup, font='Times 15 bold', exportselection=0)
    FN.focus()
    FN.grid(row=1, column=2, sticky=W)
    Label(popup, text='Enter Last Name').grid(row=2, column=1, sticky=W)
    LN = Entry(popup, font='Times 15 bold', exportselection=0)
    LN.grid(row=2, column=2, sticky=W)
    Button(popup, text='Check In', font='Times 14 bold', relief="groove", command=lambda: retrieve_name_entry()).grid(
        row=3, column=1, sticky=W)


def retrieve_name_entry():
    f = FN.get()
    l = LN.get()
    if f.strip() == '' or l.strip() == '':
        FN.delete(0, END)
        LN.delete(0, END)
    else:
        name = f + ' ' + l
        check_in(name, True)


menubar = Menu(root)
menuB = Menu(menubar, tearoff=0)
menuB2 = Menu(menubar, tearoff=0)
menuB3 = Menu(menubar, tearoff=1)
menuB.add_command(label="Exit", command=lambda: quit_handler())
menuB2.add_command(label="Attendance Graph", command=show_graph)
menuB2.add_command(label='Credit Eligible Members')
menubar.add_cascade(label="File", menu=menuB)
menubar.add_cascade(label='Analyze', menu=menuB2)
menubar.add_cascade(label='Help', menu=menuB3)
root.configure(menu=menubar)

m1 = PanedWindow(height=650, width=1000, orient=VERTICAL)
m1.pack(fill=BOTH, expand=1)

# Barcode Entry
top = Frame(m1)
m1.add(top)

barcodeEntry = Entry(top, font='Times 24 bold', width=44)
barcodeEntry.pack(side=LEFT)


def check_entry_length():
    while True:
        lengthE = len(barcodeEntry.get())
        # print(lengthE)
        try:
            if lengthE == 9 and int(lengthE) > 0:
                retrieve_entry()
        except ValueError:
            pass
        time.sleep(0.09)


thr = threading.Thread(target=check_entry_length)
thr.daemon = True
thr.start()


def retrieve_entry(name=False):
    identifier = barcodeEntry.get()
    if name is False:
        identifier = int(identifier)

    barcodeEntry.delete(0, END)
    check_in(identifier, name)


clearEntry = Button(top, text='Clear', font='Times 14 bold', relief="groove",
                    command=lambda: barcodeEntry.delete(0, END)).pack(side=LEFT)
nameEntry = Button(top, text='Name Check In', font='Times 14 bold', relief="groove", command=name_check_in_GUI)
nameEntry.pack(side=LEFT)

smallMid = Frame(m1)
m1.add(smallMid)

Label(smallMid, text='Members that have checked in', font='Times 15 bold').grid(row=1, column=1)
NumCheckIn = Label(smallMid, text="Checked In: ", font='Times 15 bold')
smallMid.grid_columnconfigure(2, minsize=1300)
NumCheckIn.grid(row=1, column=2)

middle = Frame(m1)
m1.add(middle)
# Members List
scBr = Scrollbar(middle)
scBr.pack(side=RIGHT, fill=Y)

dataList = Listbox(middle, selectmode=SINGLE, font='Times 15 bold', yscrollcommand=scBr.set)
scBr.configure(command=dataList.yview)

dataList.pack(fill=BOTH)


def add_name_listbox(row):
    global counter
    name = ws.cell(row=row, column=2).value
    name = name.split(' ')
    name = name[0][0].upper() + name[0][1:] + ' ' + name[1][0].upper() + name[1][1:]
    number = ws.cell(row=row, column=1).value
    paid = ws.cell(row=row, column=3).value
    if paid is None:
        paid = 'No'
    dataList.insert(0, name)
    signInList.insert(0, [name, number, paid, row])
    NumCheckIn.config(text='Checked In: ' + str(len(signInList)))
    dataList.selection_clear(1)
    dataList.selection_set(0)
    counter = True
    barcodeEntry.focus_force()


# Member information
bottom = Frame(m1)
m1.add(bottom)
bottom.grid_columnconfigure(3, minsize=1300)


def member_info_template(memberIndex, name='Not Applicable', number='Not Applicable', paid='No'):
    if name == paid == number and name == 'None':
        color1, color2, color3, color4 = None, None, None, None
    else:
        color1, color2, color3, color4 = None, 'lightgreen', 'lightgreen', 'red'
    eligible = "NO"
    if memberIndex != '':
        eligible = check_eligible(signInList[memberIndex][0].lower())

    if name == 'Not Applicable':
        color1 = 'Red'

    if paid == 'No':
        color3 = 'Red'

    if number == 'Not Applicable':
        color2 = 'red'

    if eligible == 'Yes':
        color4 = 'lightgreen'

    nameLab = Label(bottom, text='Name: ' + name, font='Times 15 bold', background=color1)
    nameLab.grid(row=1, column=1, sticky=W)
    numLab = Label(bottom, text='Number: ' + str(number), font='Times 15 bold', background=color2)
    numLab.grid(row=2, column=1, sticky=W)
    payLab = Label(bottom, text='Paid: ' + paid, font='Times 15 bold', background=color3)
    payLab.grid(row=3, column=1, sticky=W)
    eligLab = Label(bottom, text='Eligible for Credit: ' + eligible, font='Times 15 bold', background=color4)
    eligLab.grid(row=4, column=1, sticky=W)
    editbutton = Button(bottom, text='Edit', font='Times 13 bold', relief="groove")
    if memberIndex == '':
        editbutton.config(command=None)
    else:
        editbutton.config(
            command=lambda: edit(memberIndex, [nameLab, numLab, payLab, eligLab], editbutton, name, number, paid))
    editbutton.grid(row=1, column=3)

    return [nameLab, numLab, payLab, eligLab, editbutton]


def edit(memberIndex, delete, editbutton, name, number, paid):
    dataList.config(state='disabled')
    barcodeEntry.config(state='disabled')
    nameEntry.config(state='disabled')
    clean_mem_info_panel(delete)
    dataList.selection_set(memberIndex)
    name = name.split(' ')
    firstNameStr = name[0][0].upper() + name[0][1:]
    lastNameStr = name[1][0].upper() + name[1][1:]

    bottom.grid_columnconfigure(3, minsize=1000)
    editbutton.config(text='Done', command=lambda: save_mem_changes(memberIndex, fname, lname, sNum, payVar, editbutton,
                                                                    [radioBFrame, firstNLab, lastNLab,
                                                                     num, pay]))  # change command

    firstNLab = Label(bottom, text='First Name: ', font='Times 15 bold')
    firstNLab.grid(row=1, column=1, sticky=W)

    fname = Entry(bottom, font='Times 15 bold', exportselection=0)
    fname.insert(0, firstNameStr)
    fname.grid(row=1, column=2, sticky=W)

    lastNLab = Label(bottom, text='Last Name: ', font='Times 15 bold')
    lastNLab.grid(row=2, column=1, sticky=W)

    lname = Entry(bottom, font='Times 15 bold', exportselection=0)
    lname.grid(row=2, column=2, sticky=W)
    lname.insert(0, lastNameStr)
    num = Label(bottom, text='Number: ', font='Times 15 bold')
    num.grid(row=3, column=1, sticky=W)

    sNum = Entry(bottom, font='Times 15 bold', exportselection=0)
    sNum.insert(0, number)
    sNum.grid(row=3, column=2)

    pay = Label(bottom, text='Paid: ', font='Times 15 bold')
    pay.grid(row=4, column=1, sticky=W)
    radioBFrame = Frame(bottom)
    radioBFrame.grid(row=4, column=2, sticky=W)
    payVar = StringVar()

    Rb1 = Radiobutton(radioBFrame, text="Paid", variable=payVar, value="PAID", indicatoron=0, bd=4)
    Rb1.grid(row=1, column=1, sticky=W)
    Rb2 = Radiobutton(radioBFrame, text="Not Paid", variable=payVar, value='No', indicatoron=0, bd=4)
    Rb2.grid(row=1, column=2, sticky=W)
    if paid == 'No':
        Rb2.select()
    else:
        Rb1.select()


def save_mem_changes(memberIndex, fN, lN, num, pay, editbutton, delete):
    dataList.config(state='normal')
    barcodeEntry.config(state='normal')
    nameEntry.config(state='normal')
    bottom.grid_columnconfigure(3, minsize=1300)
    barcodeEntry.focus_force()
    dataList.selection_clear(0, END)
    name = fN.get() + ' ' + lN.get()
    name = name.lower()
    number = num.get()
    try:
        number = int(number)
    except ValueError:
        num.delete(0, END)
        return None
    payment = pay.get()
    if payment == 'No':
        payment = None
    row = signInList[memberIndex][3]
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=1, value=number)
    ws.cell(row=row, column=3, value=payment)
    #(signInList[memberIndex])
    name = name.split(' ')
    name = name[0][0].upper() + name[0][1:] + ' ' + name[1][0].upper() + name[1][1:]
    if payment is None:
        payment = 'No'
    signInList[memberIndex] = [name, number, payment, signInList[memberIndex][-1]]
    #print(signInList[memberIndex])
    dataList.delete(memberIndex)
    dataList.insert(memberIndex, name)
    save()
    delete.append(fN)
    delete.append(lN)
    delete.append(num)
    clean_mem_info_panel(delete)
    nameLab, numLab, payLab, eligLab, paid = 0, 0, 0, 0, 0
    editbutton.config(text='Edit',
                      command=lambda: edit([nameLab, numLab, payLab, eligLab], editbutton, name, number, paid))


def display_member_info(memberIndex):
    memberIndex = (str(memberIndex))
    memberIndex = memberIndex.replace('(', '')
    memberIndex = memberIndex.replace(')', '')
    memberIndex = memberIndex.replace(',', '')
    if memberIndex is '':
        labelIDs = member_info_template(memberIndex, 'None', 'None', 'None')

    else:
        memberIndex = int(memberIndex)
        labelIDs = member_info_template(memberIndex, signInList[memberIndex][0], signInList[memberIndex][1],
                                        signInList[memberIndex][2])
    return labelIDs


def clean_mem_info_panel(labelIDList):
    if labelIDList is None:
        pass
    else:
        for i in labelIDList:
            i.destroy()


def check_selection():
    global counter
    b = None
    lastPos = 0
    while True:
        curSelect = dataList.curselection()
        if counter or curSelect != lastPos:
            clean_mem_info_panel(b)
            b = display_member_info(curSelect)
            counter = False
        time.sleep(0.1)
        lastPos = curSelect


listboxThr = threading.Thread(target=check_selection)
listboxThr.daemon = True
listboxThr.start()


def quit_handler():
    root.destroy()
    Stp()


barcodeEntry.focus_force()
root.protocol("WM_DELETE_WINDOW", quit_handler)
root.mainloop()
